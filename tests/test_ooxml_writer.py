"""Tests de non-régression du moteur OOXML."""

from pathlib import Path
import zipfile

import pandas as pd
from lxml import etree
from openpyxl import load_workbook

from ooxml_writer import NS_MAIN, OoxmlWorkbook
from proposition_formulas import write_proposition_formulas


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "Analyse_MASK.xlsx"


def test_ecriture_preserve_le_modele_et_force_le_recalcul(tmp_path):
    """Le moteur doit écrire les données sans reconstruire le classeur."""
    output = tmp_path / "rapport.xlsx"
    workbook = OoxmlWorkbook(TEMPLATE)
    workbook.sheets["Analyse"].range("D1").value = "Ellebore"
    data = workbook.sheets["DATA"]
    data.used_range.clear_contents()
    data.range("A1").options(index=False, header=True).value = pd.DataFrame(
        {"ISIN": ["TEST"], "Poids": [0.25]}
    )
    write_proposition_formulas(
        workbook.sheets["Proposition"],
        data_rows=2,
        fund_rows=2,
        benchmark_rows=2,
    )
    workbook.save(output)

    with zipfile.ZipFile(TEMPLATE) as source, zipfile.ZipFile(output) as result:
        assert result.testzip() is None
        assert result.read("xl/charts/chart1.xml") == source.read(
            "xl/charts/chart1.xml"
        )
        assert "xl/calcChain.xml" not in result.namelist()
        root = etree.fromstring(result.read("xl/workbook.xml"))
        calc = root.find(f"{{{NS_MAIN}}}calcPr")
        assert calc.get("calcMode") == "auto"
        assert calc.get("fullCalcOnLoad") == "1"
        assert calc.get("forceFullCalc") == "1"

    book = load_workbook(output, read_only=True, data_only=False)
    assert book["Analyse"]["D1"].value == "Ellebore"
    assert book["DATA"]["A2"].value == "TEST"
    assert book["DATA"]["B2"].value == 0.25
    assert book["Proposition"]["A6"].value.startswith("=IF(")
    assert "'DATA'!" in book["Proposition"]["B6"].value
    assert "_xlfn." not in book["Proposition"]["A6"].value
    book.close()
